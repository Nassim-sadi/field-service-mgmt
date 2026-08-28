import csv
import io

from django.db import transaction
from django.db.models import Avg, Count, Sum
from django.http import StreamingHttpResponse
from django.utils import timezone
from rest_framework import permissions, status, viewsets
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response

from .models import (
    Asset,
    AuditLog,
    Company,
    Customer,
    Part,
    ServiceReport,
    Site,
    Technician,
    TimeEntry,
    WorkOrder,
    WorkOrderPart,
)
from .permissions import IsManagement
from .serializers import (
    AssetSerializer,
    CompanySerializer,
    CustomerSerializer,
    PartSerializer,
    ServiceReportSerializer,
    SiteSerializer,
    TechnicianSerializer,
    TimeEntrySerializer,
    WorkOrderPartSerializer,
    WorkOrderSerializer,
)
from .services import perform_transition


class CompanyViewSet(viewsets.ModelViewSet):
    queryset = Company.objects.all()
    serializer_class = CompanySerializer


class CustomerViewSet(viewsets.ModelViewSet):
    queryset = Customer.objects.select_related("company").all()
    serializer_class = CustomerSerializer
    filterset_fields = ["company"]
    search_fields = ["name", "email", "phone"]

    @action(detail=True)
    def sites(self, request, pk=None):
        customer = self.get_object()
        serializer = SiteSerializer(customer.sites.all(), many=True)
        return Response(serializer.data)

    @action(detail=False, methods=["post"], parser_classes=[MultiPartParser], permission_classes=[IsManagement])
    def import_csv(self, request):
        if "netlify.app" in request.META.get("HTTP_ORIGIN", "") or "netlify.app" in request.META.get("HTTP_REFERER", ""):
            return Response({"detail": "Demo — import disabled on Netlify"}, status=status.HTTP_403_FORBIDDEN)
        file = request.FILES.get("file")
        if not file:
            return Response({"detail": "file is required"}, status=status.HTTP_400_BAD_REQUEST)
        on_duplicate = request.data.get("on_duplicate", "skip")
        if on_duplicate not in ("skip", "overwrite"):
            on_duplicate = "skip"
        name = file.name.lower()
        is_csv = name.endswith(".csv")
        is_xlsx = name.endswith(".xlsx") or name.endswith(".xls")
        if not (is_csv or is_xlsx):
            return Response({"detail": "Only .csv and .xlsx supported"}, status=status.HTTP_400_BAD_REQUEST)

        def normalize(h: str) -> str:
            h = h.strip().lower()
            aliases = {"société": "company", "societe": "company", "téléphone": "phone", "telephone": "phone", "adresse": "address", "nom": "name"}
            return aliases.get(h, h)

        expected = {"company", "name", "email"}
        headers: list[str] = []
        rows_iter = None
        errors: list[dict] = []

        try:
            if is_csv:
                text = io.TextIOWrapper(file, encoding="utf-8-sig")
                reader = csv.DictReader(text)
                if reader.fieldnames is None:
                    return Response({"detail": "Empty file or missing header"}, status=status.HTTP_400_BAD_REQUEST)
                headers = [normalize(h) for h in reader.fieldnames]
                if not expected.issubset(set(headers)):
                    return Response({"detail": f"Missing headers: {', '.join(expected - set(headers))}", "expected": sorted(expected)}, status=status.HTTP_400_BAD_REQUEST)

                def gen_csv():
                    for idx, row in enumerate(reader, start=2):
                        norm = {normalize(k): (v.strip() if isinstance(v, str) else v) for k, v in row.items()}
                        yield idx, norm

                rows_iter = gen_csv()
            else:
                import openpyxl

                wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
                ws = wb.active
                it = ws.iter_rows(values_only=True)
                try:
                    raw_headers = next(it)
                except StopIteration:
                    return Response({"detail": "Empty file or missing header"}, status=status.HTTP_400_BAD_REQUEST)
                headers = [normalize(str(h) if h is not None else "") for h in raw_headers]
                if not expected.issubset(set(headers)):
                    return Response({"detail": f"Missing headers: {', '.join(expected - set(headers))}", "expected": sorted(expected)}, status=status.HTTP_400_BAD_REQUEST)

                def gen_xlsx():
                    for idx, row in enumerate(it, start=2):
                        norm = {headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row) if i < len(headers)}
                        yield idx, norm

                rows_iter = gen_xlsx()
        except Exception as e:
            return Response({"detail": f"Failed to read file: {e}"}, status=status.HTTP_400_BAD_REQUEST)

        company_cache = {c.name.lower(): c for c in Company.objects.all()}
        existing = {(c.email.lower(), c.company_id): c for c in Customer.objects.select_related("company").all()}
        to_create: list[Customer] = []
        to_update: list[Customer] = []
        imported = skipped = overwritten = 0

        for idx, row in rows_iter:  # type: ignore
            company_name = row.get("company", "").strip()
            name = row.get("name", "").strip()
            email = row.get("email", "").strip()
            if not company_name or not name or not email:
                errors.append({"row": idx, "field": "company/name/email", "message": "company, name, email required"})
                continue
            company = company_cache.get(company_name.lower())
            if not company:
                company = Company.objects.create(name=company_name)
                company_cache[company_name.lower()] = company
            key = (email.lower(), company.id)
            if key in existing:
                if on_duplicate == "skip":
                    skipped += 1
                    continue
                cust = existing[key]
                cust.name = name
                cust.phone = row.get("phone", "")[:40]
                cust.address = row.get("address", "")
                to_update.append(cust)
                overwritten += 1
            else:
                cust = Customer(company=company, name=name, email=email, phone=row.get("phone", "")[:40], address=row.get("address", ""))
                to_create.append(cust)
                existing[key] = cust

        with transaction.atomic():
            if to_create:
                Customer.objects.bulk_create(to_create, batch_size=500)
                imported = len(to_create)
            if to_update:
                Customer.objects.bulk_update(to_update, ["name", "phone", "address"], batch_size=500)

        return Response({"imported": imported, "overwritten": overwritten, "skipped": skipped, "failed": len(errors), "errors": errors[:100]})

    @action(detail=False, methods=["get"], permission_classes=[IsManagement])
    def export(self, request):
        fmt = request.query_params.get("format", "csv")
        qs = self.filter_queryset(self.get_queryset()).select_related("company").only("name", "email", "phone", "address", "company__name").iterator(chunk_size=500)

        if fmt == "xlsx":
            import openpyxl
            from django.http import HttpResponse

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "customers"
            ws.append(["company", "name", "email", "phone", "address"])
            for c in qs:
                ws.append([c.company.name, c.name, c.email, c.phone, c.address])
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            resp = HttpResponse(buf.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            resp["Content-Disposition"] = 'attachment; filename="customers.xlsx"'
            return resp

        def gen():
            yield "company,name,email,phone,address\n"
            for c in qs:
                def esc(v: str) -> str:
                    v = (v or "").replace('"', '""')
                    return f'"{v}"' if "," in v or '"' in v or "\n" in v else v
                yield f"{esc(c.company.name)},{esc(c.name)},{esc(c.email)},{esc(c.phone)},{esc(c.address)}\n"

        resp = StreamingHttpResponse(gen(), content_type="text/csv")
        resp["Content-Disposition"] = 'attachment; filename="customers.csv"'
        return resp


class SiteViewSet(viewsets.ModelViewSet):
    queryset = Site.objects.select_related("customer").all()
    serializer_class = SiteSerializer
    filterset_fields = ["customer"]
    search_fields = ["name", "address"]

    @action(detail=True)
    def assets(self, request, pk=None):
        site = self.get_object()
        serializer = AssetSerializer(site.assets.all(), many=True)
        return Response(serializer.data)


class AssetViewSet(viewsets.ModelViewSet):
    queryset = Asset.objects.select_related("site").all()
    serializer_class = AssetSerializer
    filterset_fields = ["site", "status"]
    search_fields = ["name", "serial_number", "asset_type"]


class TechnicianViewSet(viewsets.ModelViewSet):
    queryset = Technician.objects.select_related("user").all()
    serializer_class = TechnicianSerializer
    filterset_fields = ["is_active", "specialty"]
    search_fields = ["user__username", "user__first_name", "user__last_name", "specialty"]


class PartViewSet(viewsets.ModelViewSet):
    queryset = Part.objects.all()
    serializer_class = PartSerializer
    search_fields = ["sku", "name"]


class WorkOrderPartViewSet(viewsets.ModelViewSet):
    queryset = WorkOrderPart.objects.select_related("part", "work_order").all()
    serializer_class = WorkOrderPartSerializer
    filterset_fields = ["work_order"]


class TimeEntryViewSet(viewsets.ModelViewSet):
    queryset = TimeEntry.objects.select_related("work_order", "technician").all()
    serializer_class = TimeEntrySerializer
    filterset_fields = ["work_order", "technician"]


class ServiceReportViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = ServiceReport.objects.select_related("work_order").all()
    serializer_class = ServiceReportSerializer
    filterset_fields = ["work_order"]


class WorkOrderViewSet(viewsets.ModelViewSet):
    queryset = (
        WorkOrder.objects.select_related("customer", "site", "asset", "assigned_technician")
        .prefetch_related("parts", "parts__part", "time_entries", "files", "audit_logs")
        .all()
    )
    serializer_class = WorkOrderSerializer
    filterset_fields = ["status", "priority", "assigned_technician", "site", "customer"]
    search_fields = ["number", "title", "description", "customer__name"]

    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user
        if user.is_authenticated and user.is_technician:
            qs = qs.filter(assigned_technician__user=user)
        elif user.is_authenticated and user.role == "customer":
            qs = qs.filter(customer__user=user)
        return qs

    def _transition(self, request, pk, target):
        work_order = self.get_object()
        instance = perform_transition(work_order, target, user=request.user)
        return Response(self.get_serializer(instance).data)

    @action(detail=True, methods=["post"], permission_classes=[IsManagement])
    def assign(self, request, pk=None):
        work_order = self.get_object()
        technician_id = request.data.get("technician")
        if not technician_id:
            return Response({"detail": "technician is required."}, status=status.HTTP_400_BAD_REQUEST)
        try:
            work_order.assigned_technician = Technician.objects.get(pk=technician_id)
        except Technician.DoesNotExist:
            return Response({"detail": "technician not found."}, status=status.HTTP_400_BAD_REQUEST)
        instance = perform_transition(work_order, WorkOrder.Status.ASSIGNED, user=request.user)
        return Response(self.get_serializer(instance).data)

    @action(detail=True, methods=["post"], permission_classes=[permissions.IsAuthenticated])
    def accept(self, request, pk=None):
        return self._transition(request, pk, WorkOrder.Status.ACCEPTED)

    @action(detail=True, methods=["post"], permission_classes=[permissions.IsAuthenticated])
    def start(self, request, pk=None):
        return self._transition(request, pk, WorkOrder.Status.IN_PROGRESS)

    @action(detail=True, methods=["post"], permission_classes=[permissions.IsAuthenticated])
    def complete(self, request, pk=None):
        work_order = self.get_object()
        instance = perform_transition(work_order, WorkOrder.Status.COMPLETED, user=request.user)
        report_data = request.data
        if report_data:
            serializer = ServiceReportSerializer(data={**report_data, "work_order": work_order.id})
            serializer.is_valid(raise_exception=True)
            serializer.save()
        return Response(self.get_serializer(instance).data)

    @action(detail=True, methods=["post"], permission_classes=[IsManagement])
    def reopen(self, request, pk=None):
        return self._transition(request, pk, WorkOrder.Status.NEW)


class DashboardViewSet(viewsets.GenericViewSet):
    permission_classes = [permissions.IsAuthenticated]

    @action(detail=False)
    def stats(self, request):
        open_count = WorkOrder.objects.filter(
            status__in=[
                WorkOrder.Status.NEW,
                WorkOrder.Status.ASSIGNED,
                WorkOrder.Status.ACCEPTED,
                WorkOrder.Status.IN_PROGRESS,
            ]
        ).count()
        in_progress = WorkOrder.objects.filter(status=WorkOrder.Status.IN_PROGRESS).count()
        completed = WorkOrder.objects.filter(status=WorkOrder.Status.COMPLETED).count()
        technician_count = Technician.objects.filter(is_active=True).count()
        overdue = WorkOrder.objects.filter(
            status__in=[
                WorkOrder.Status.NEW,
                WorkOrder.Status.ASSIGNED,
                WorkOrder.Status.ACCEPTED,
                WorkOrder.Status.IN_PROGRESS,
            ],
            due_at__isnull=False,
        ).filter(due_at__lt=timezone.now()).count()
        return Response(
            {
                "open_tickets": open_count,
                "in_progress": in_progress,
                "completed": completed,
                "technicians": technician_count,
                "overdue": overdue,
            }
        )

    @action(detail=False)
    def workload(self, request):
        rows = (
            WorkOrder.objects.exclude(
                status__in=[WorkOrder.Status.COMPLETED, WorkOrder.Status.CANCELLED]
            )
            .values("assigned_technician__user__first_name", "assigned_technician__user__last_name")
            .annotate(count=Count("id"))
            .order_by("-count")
        )
        return Response(
            [
                {
                    "technician": f"{row['assigned_technician__user__first_name']} {row['assigned_technician__user__last_name']}".strip(),
                    "count": row["count"],
                }
                for row in rows
            ]
        )

    @action(detail=False)
    def sla(self, request):
        overdue = WorkOrder.objects.filter(
            status__in=[
                WorkOrder.Status.NEW,
                WorkOrder.Status.ASSIGNED,
                WorkOrder.Status.ACCEPTED,
                WorkOrder.Status.IN_PROGRESS,
            ],
            due_at__isnull=False,
        ).filter(
            due_at__lt=timezone.now()
        )
        total_open = WorkOrder.objects.filter(
            status__in=[
                WorkOrder.Status.NEW,
                WorkOrder.Status.ASSIGNED,
                WorkOrder.Status.ACCEPTED,
                WorkOrder.Status.IN_PROGRESS,
            ]
        ).count()
        return Response(
            {
                "breaches": overdue.count(),
                "open_orders": total_open,
                "on_time": max(0, total_open - overdue.count()),
            }
        )

    @action(detail=False)
    def avg_resolution(self, request):
        agg = WorkOrder.objects.filter(status=WorkOrder.Status.COMPLETED).aggregate(
            avg=Avg("resolution_minutes")
        )
        return Response({"avg_resolution_minutes": agg["avg"]})

    @action(detail=False)
    def parts_consumption(self, request):
        rows = (
            WorkOrderPart.objects.values("part__name", "part__sku")
            .annotate(total_qty=Sum("quantity"))
            .order_by("-total_qty")
        )
        return Response(
            [
                {"part": row["part__name"], "sku": row["part__sku"], "quantity": row["total_qty"]}
                for row in rows
            ]
        )

    @action(detail=False)
    def over_time(self, request):
        rows = (
            WorkOrder.objects.exclude(status__in=[WorkOrder.Status.NEW])
            .values("status")
            .annotate(count=Count("id"))
        )
        return Response([{"status": row["status"], "count": row["count"]} for row in rows])
